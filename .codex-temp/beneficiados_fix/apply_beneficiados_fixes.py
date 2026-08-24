from __future__ import annotations

import json
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


INPUT_PATH = Path(r"C:\Users\USUARIO\Downloads\SEGUIMIENTO_CORTE_A_CORTE_2026-1_CONSOLIDADO_1Y2.xlsx")
PATCH_PATH = Path(r"C:\xampp\htdocs\Bienestar\storage\app\exports\beneficiados_123_fixes.json")
OUTPUT_PATH = Path(r"C:\xampp\htdocs\Bienestar\storage\app\exports\SEGUIMIENTO_CORTE_A_CORTE_2026-1_CONSOLIDADO_1Y2_corregido.xlsx")
SHEET_NAME = "BENEFICIADOS 1,2Y3CORTE"


def norm_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


def canonical(value: object) -> str:
    text = norm_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("�", "")
    text = " ".join(text.split())
    return text.upper()


def main() -> None:
    patches = json.loads(PATCH_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(INPUT_PATH)
    worksheet = workbook[SHEET_NAME]

    header_map: dict[str, int] = {}
    for cell in worksheet[1]:
        header = canonical(cell.value)
        if header:
            header_map[header] = cell.column

    for patch in patches:
        row = int(patch["row"])
        field = canonical(patch["field"])
        value = patch["value"]
        column = header_map.get(field)
        if column is None:
            continue
        worksheet.cell(row=row, column=column, value=value)

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
