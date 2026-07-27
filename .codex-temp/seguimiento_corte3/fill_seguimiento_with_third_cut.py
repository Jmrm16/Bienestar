from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook


CORRECTED_PATH = Path(r"C:\Users\USUARIO\Downloads\FORMATO_SEGUIMIENTO_NOTAS_CORREGIDO_COMPLETADO.xlsx")
THIRD_CUT_PATH = Path(r"C:\Users\USUARIO\Downloads\Notas III CORTE 2026-I.xlsx")
OUTPUT_PATH = Path(r"C:\xampp\htdocs\Bienestar\storage\app\exports\SEGUIMIENTO_CORTE_A_CORTE_2026-1_CONSOLIDADO_1Y2Y3.xlsx")


def norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def maybe_number(value):
    if value in ("", None):
        return None
    try:
        return float(value)
    except Exception:
        return value


source_wb = load_workbook(THIRD_CUT_PATH, data_only=True)
source_ws = source_wb[source_wb.sheetnames[0]]

source_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in source_ws[1]]
source_index = {header: idx for idx, header in enumerate(source_headers)}

source_map = {}
for row in source_ws.iter_rows(min_row=2, values_only=True):
    key = (
        norm(row[source_index["Identificacion"]]),
        norm(row[source_index["Ide Materia"]]),
        norm(row[source_index["Grupo"]]),
    )
    if key == ("", "", ""):
        continue
    source_map[key] = {
        "nota1": maybe_number(row[source_index["1er"]]),
        "nota2": maybe_number(row[source_index["2er"]]),
        "nota3": maybe_number(row[source_index["3er"]]),
        "final": maybe_number(row[source_index["Final"]]),
    }

target_wb = load_workbook(CORRECTED_PATH)
target_ws = target_wb["SEGUIMIENTO"]

target_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in target_ws[1]]
target_index = {header: idx + 1 for idx, header in enumerate(target_headers)}

filled_rows = 0
matched_rows = 0
for row_idx in range(2, target_ws.max_row + 1):
    key = (
        norm(target_ws.cell(row=row_idx, column=target_index["Identificación"]).value),
        norm(target_ws.cell(row=row_idx, column=target_index["CODIGO DE LA MATERIA"]).value),
        norm(target_ws.cell(row=row_idx, column=target_index["Grupo"]).value),
    )
    source = source_map.get(key)
    if not source:
        continue
    matched_rows += 1

    current_values = [
        target_ws.cell(row=row_idx, column=target_index["Notas 1er corte"]).value,
        target_ws.cell(row=row_idx, column=target_index["Notas 2do corte"]).value,
        target_ws.cell(row=row_idx, column=target_index["Nota 3er corte"]).value,
        target_ws.cell(row=row_idx, column=target_index["Final"]).value,
    ]
    next_values = [source["nota1"], source["nota2"], source["nota3"], source["final"]]

    if [None if value == "" else value for value in current_values] != next_values:
        filled_rows += 1

    target_ws.cell(row=row_idx, column=target_index["Notas 1er corte"]).value = source["nota1"]
    target_ws.cell(row=row_idx, column=target_index["Notas 2do corte"]).value = source["nota2"]
    target_ws.cell(row=row_idx, column=target_index["Nota 3er corte"]).value = source["nota3"]
    target_ws.cell(row=row_idx, column=target_index["Final"]).value = source["final"]

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
target_wb.save(OUTPUT_PATH)

print(
    {
        "output": str(OUTPUT_PATH),
        "matched_rows": matched_rows,
        "updated_rows": filled_rows,
        "total_rows": target_ws.max_row - 1,
    }
)
