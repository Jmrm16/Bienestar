from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


CONSOLIDATED_PATH = Path(r"C:\Users\USUARIO\Downloads\SEGUIMIENTO_CORTE_A_CORTE_2026-1_CONSOLIDADO_1Y2.xlsx")
CORRECTED_PATH = Path(r"C:\Users\USUARIO\Downloads\FORMATO_SEGUIMIENTO_NOTAS_CORREGIDO_COMPLETADO.xlsx")
OUTPUT_PATH = Path(r"C:\xampp\htdocs\Bienestar\storage\app\exports\SEGUIMIENTO_CORTE_A_CORTE_2026-1_CONSOLIDADO_1Y2Y3.xlsx")


def copy_sheet_contents(source_ws, target_ws) -> None:
    # remove existing rows/cols content
    if target_ws.max_row > 1:
        target_ws.delete_rows(1, target_ws.max_row)
    if target_ws.max_column > 1:
        target_ws.delete_cols(1, target_ws.max_column)

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
            if cell.protection:
                new_cell.protection = copy(cell.protection)

    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
        target_ws.column_dimensions[key].hidden = dim.hidden
        target_ws.column_dimensions[key].bestFit = dim.bestFit
        target_ws.column_dimensions[key].outlineLevel = dim.outlineLevel

    for idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx].height = dim.height
        target_ws.row_dimensions[idx].hidden = dim.hidden
        target_ws.row_dimensions[idx].outlineLevel = dim.outlineLevel

    for merged_range in list(source_ws.merged_cells.ranges):
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    target_ws.sheet_properties.tabColor = copy(source_ws.sheet_properties.tabColor)


consolidated_wb = load_workbook(CONSOLIDATED_PATH)
corrected_wb = load_workbook(CORRECTED_PATH)

source_ws = corrected_wb["SEGUIMIENTO"]
target_ws = consolidated_wb.worksheets[0]
target_ws.title = "SEGUIMIENTO"

copy_sheet_contents(source_ws, target_ws)

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
consolidated_wb.save(OUTPUT_PATH)

print({
    "output": str(OUTPUT_PATH),
    "first_sheet_title": consolidated_wb.worksheets[0].title,
    "first_sheet_rows": consolidated_wb.worksheets[0].max_row,
    "first_sheet_cols": consolidated_wb.worksheets[0].max_column,
    "all_sheets": consolidated_wb.sheetnames,
})
